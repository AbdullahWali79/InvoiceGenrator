"""Excel repository for loading and updating medicine inventory."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app import config
from app.models.medicine import Medicine


@dataclass
class ExcelColumnMap:
    """Column indexes for required fields."""

    name: int
    mg: int
    rack_number: int
    stock: int
    price: int
    actual_price: int


class ExcelRepository:
    """Handles reading and writing inventory data in Excel."""

    def __init__(self, path: Path | str = None, sheet_name: Optional[str] = None) -> None:
        self.path: Path = Path(path) if path else config.EXCEL_PATH
        self.sheet_name = sheet_name or config.EXCEL_SHEET_NAME
        self._workbook = load_workbook(self.path)
        self._sheet: Worksheet = self._workbook[self.sheet_name]
        self._columns = self._detect_columns()

    def _detect_columns(self) -> ExcelColumnMap:
        """Detect columns from header row; raises if missing."""
        headers: Dict[str, int] = {}
        for idx, cell in enumerate(self._sheet[1], start=1):
            headers[str(cell.value).strip()] = idx

        required = ["Medicine_Name", "MG", "Rack_Number", "Stock", "Price", "Actual_Price"]
        missing = [col for col in required if col not in headers]
        if missing:
            raise ValueError(f"Missing required columns in Excel: {', '.join(missing)}")

        return ExcelColumnMap(
            name=headers["Medicine_Name"],
            mg=headers["MG"],
            rack_number=headers["Rack_Number"],
            stock=headers["Stock"],
            price=headers["Price"],
            actual_price=headers["Actual_Price"],
        )

    def list_medicines(self) -> List[Medicine]:
        """Return all medicines as dataclasses."""
        medicines: List[Medicine] = []
        for row in self._sheet.iter_rows(min_row=2, values_only=False):
            name = row[self._columns.name - 1].value
            if not name:
                continue

            mg = row[self._columns.mg - 1].value or ""
            rack_number = row[self._columns.rack_number - 1].value or ""
            stock = self._to_int(row[self._columns.stock - 1].value, default=0)
            price = self._to_float(row[self._columns.price - 1].value, default=0.0)
            actual_price = self._to_float(row[self._columns.actual_price - 1].value, default=None)

            medicines.append(
                Medicine(
                    name=str(name),
                    mg=str(mg),
                    rack_number=str(rack_number),
                    stock=stock,
                    price=price,
                    actual_price=actual_price,
                )
            )
        return medicines

    def save_stock_updates(self, updated_stocks: Dict[str, int]) -> None:
        """Persist stock updates back to Excel."""
        for row in self._sheet.iter_rows(min_row=2):
            name_cell = row[self._columns.name - 1]
            stock_cell = row[self._columns.stock - 1]
            if name_cell.value in updated_stocks:
                stock_cell.value = updated_stocks[name_cell.value]
        self._workbook.save(self.path)

    @staticmethod
    def _to_float(value, default: Optional[float]) -> Optional[float]:
        if value in (None, ""):
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _to_int(value, default: int) -> int:
        if value in (None, ""):
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default
