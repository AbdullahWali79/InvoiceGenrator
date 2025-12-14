"""Excel-backed data store for pharmacy inventory."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app import config


REQUIRED_COLUMNS = ["Medicine_Name", "Price", "Stock", "MG", "Rack_Number", "Actual_Price"]


def _normalize_name(name: str) -> str:
    return str(name).strip().lower()


@dataclass
class _RowRef:
    row_index: int
    stock_cell_ref: str


class ExcelDataStore:
    """Loads and mutates inventory from an Excel sheet."""

    def __init__(self, path: Path | str = None, sheet_name: str | None = None) -> None:
        self.path: Path = Path(path) if path else config.EXCEL_PATH
        self.sheet_name = sheet_name or config.EXCEL_SHEET_NAME
        self._workbook = None
        self._sheet: Optional[Worksheet] = None
        self._col_map: Dict[str, int] = {}
        self._names: List[str] = []
        self._records: Dict[str, Dict] = {}
        self._row_refs: Dict[str, _RowRef] = {}
        self._load()

    def _load(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")

        self._workbook = load_workbook(self.path)
        if self.sheet_name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found in Excel file.")

        self._sheet = self._workbook[self.sheet_name]
        self._col_map = self._detect_columns()
        self._read_rows()

    def _detect_columns(self) -> Dict[str, int]:
        headers: Dict[str, int] = {}
        for idx, cell in enumerate(self._sheet[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = idx

        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        return headers

    def _read_rows(self) -> None:
        self._names.clear()
        self._records.clear()
        self._row_refs.clear()

        for row_idx, row in enumerate(self._sheet.iter_rows(min_row=2), start=2):
            name_val = row[self._col_map["Medicine_Name"] - 1].value
            if name_val in (None, ""):
                continue

            mg = row[self._col_map["MG"] - 1].value or ""
            rack = row[self._col_map["Rack_Number"] - 1].value or ""
            stock = self._to_int(row[self._col_map["Stock"] - 1].value, default=0)
            price_val = self._to_float(row[self._col_map["Price"] - 1].value, default=0.0)
            actual_price_val = self._to_float(
                row[self._col_map["Actual_Price"] - 1].value, default=None
            )
            effective_price = (
                actual_price_val if actual_price_val not in (None, "") else price_val
            )

            norm_name = _normalize_name(name_val)
            record = {
                "name": str(name_val),
                "mg": str(mg),
                "rack": str(rack),
                "stock": stock,
                "price": effective_price,
            }
            self._names.append(record["name"])
            self._records[norm_name] = record
            stock_cell = row[self._col_map["Stock"] - 1].coordinate
            self._row_refs[norm_name] = _RowRef(row_index=row_idx, stock_cell_ref=stock_cell)

    @staticmethod
    def _to_int(value, default: int) -> int:
        if value in (None, ""):
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _to_float(value, default: Optional[float]) -> Optional[float]:
        if value in (None, ""):
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def get_all_names(self) -> List[str]:
        """Return display names in sheet order."""
        return list(self._names)

    def get_medicine(self, name: str) -> Optional[Dict]:
        """Return a record dict or None if not found."""
        return self._records.get(_normalize_name(name))

    def reduce_stock(self, name: str, qty: int) -> None:
        """Reduce stock for a medicine. Raises if insufficient."""
        if qty <= 0:
            raise ValueError("Quantity must be positive.")
        norm = _normalize_name(name)
        if norm not in self._records:
            raise KeyError(f"Medicine '{name}' not found.")

        record = self._records[norm]
        new_stock = record["stock"] - qty
        if new_stock < 0:
            raise ValueError(
                f"Insufficient stock for '{record['name']}'. "
                f"Available: {record['stock']}, requested: {qty}."
            )

        record["stock"] = new_stock
        stock_cell_ref = self._row_refs[norm].stock_cell_ref
        self._sheet[stock_cell_ref].value = new_stock

    def save(self) -> None:
        """Persist changes to disk."""
        self._workbook.save(self.path)

    def self_check(self) -> bool:
        """Verify required columns exist; returns True when valid."""
        return all(col in self._col_map for col in REQUIRED_COLUMNS)

